#!/usr/bin/env python3
"""
Populate an Event Panel worksheet from Zoom exports.

Example:
    python generate_panel_sheet.py \
        --workbook "Event_Panel/Events_ PANELS.xlsx" \
        --template "TEMPLATE(MAKE A COPY)" \
        --sheet-name "23 Oct 2025 - VET" \
        --product VET \
        --registration "Event_Panel/81314758372 - Registration Report.csv" \
        --attendee "Event_Panel/81314758372 - Attendee Report.csv" \
        --survey "Event_Panel/report_survey (2).csv" \
        --output "Event_Panel/Events_ PANELS_UPDATED.xlsx"
"""

from __future__ import annotations

import argparse
import csv
import datetime as dt
import warnings
from collections import defaultdict
from pathlib import Path
from typing import Dict, List, Optional, Tuple

from openpyxl import load_workbook

warnings.filterwarnings("ignore", category=UserWarning)

# ICP heuristics – tune as the sales team refines their criteria.
ALLOWED_COUNTRIES = {
    "United States",
    "Canada",
    "United Kingdom",
    "Australia",
    "New Zealand",
    "United Arab Emirates",
}

ICP_POSITIVE = [
    "practice manager",
    "office manager",
    "practice owner",
    "owner",
    "consultant",
    "coach",
]

ICP_NEGATIVE = [
    "assistant professor",
    "nurse",
    "technician",
    "employee",
    "dentist",
    "engineer",
    "physician extender",
    "notetaker",
]

ICP_REVIEW = [
    "ceo",
    "founder",
    "president",
    "director",
    "cso",
    "veterinarian",
]

PARTNER_KEYWORDS = {
    "rhonda bell",
    "amanda landis-hanna",
    "karen",
    "partner",
}


def load_registration_report(
    path: Path,
) -> Tuple[Dict[str, str], List[Dict[str, str]]]:
    """Read Zoom registration report and return metadata + registrants."""
    rows = list(csv.reader(path.open(newline="", encoding="utf-8-sig")))
    metadata: Dict[str, str] = {}
    registrants: List[Dict[str, str]] = []
    header: Optional[List[str]] = None

    for idx, row in enumerate(rows):
        if not row:
            continue
        first = row[0].strip()
        if first == "Topic":
            next_row = rows[idx + 1]
            metadata = {
                rows[idx][col].strip(): next_row[col].strip()
                for col in range(len(rows[idx]))
                if rows[idx][col].strip()
            }
        if first == "First Name":
            header = row
            start_idx = idx + 1
            break

    if header is None:
        raise ValueError("Could not locate registration table header.")

    for row in rows[start_idx:]:
        if not any(row):
            continue
        if row[0].strip() == "Attendee Details":
            break
        if len(row) < len(header):
            row += [""] * (len(header) - len(row))
        registrants.append(
            {header[col]: row[col].strip() for col in range(len(header))}
        )

    return metadata, registrants


def load_attendance(path: Path) -> Dict[str, Dict[str, object]]:
    """Aggregate Zoom attendance records by email."""
    rows = list(csv.reader(path.open(newline="", encoding="utf-8-sig")))
    header: Optional[List[str]] = None
    start_idx = 0

    for idx, row in enumerate(rows):
        if row and row[0].strip() == "Attendee Details":
            header = rows[idx + 1]
            start_idx = idx + 2
            break

    if header is None:
        raise ValueError("Could not locate attendee table header.")

    records: Dict[str, Dict[str, object]] = defaultdict(
        lambda: {"attended": False, "minutes": 0}
    )

    for row in rows[start_idx:]:
        if not any(row):
            continue
        label = row[0].strip()
        if label in {"Host Details", "Panelist Details"}:
            break
        if len(row) < len(header):
            row += [""] * (len(header) - len(row))

        entry = {header[col]: row[col].strip() for col in range(len(header))}
        email = entry.get("Email", "").lower()
        if not email:
            continue

        attended = entry.get("Attended", "").lower() == "yes"
        minutes_raw = entry.get("Time in Session (minutes)", "")
        minutes = int(minutes_raw) if minutes_raw.isdigit() else 0

        records[email]["attended"] = records[email]["attended"] or attended
        records[email]["minutes"] = int(records[email]["minutes"]) + minutes

    return records


def classify_icp(job_title: str, country: str) -> Tuple[str, str]:
    """Return ICP classification and notes based on heuristics."""
    job = (job_title or "").lower()
    country_name = (country or "").title()

    if country_name not in ALLOWED_COUNTRIES:
        return "Non-ICP", "country outside ICP list"

    if any(term in job for term in ICP_NEGATIVE):
        return "Non-ICP", "role excluded by guideline"

    if any(term in job for term in ICP_POSITIVE):
        return "ICP Confirmed", "matches core ICP role"

    if any(term in job for term in ICP_REVIEW):
        if "owner" in job:
            return "ICP Confirmed", "veterinarian owner"
        return "Needs Review", "leadership or clinician – confirm ownership"

    if "owner" in job:
        return "ICP Confirmed", "owner role"

    return "Needs Review", "no matching keyword – manual check"


def derive_lead_type(source_name: str) -> str:
    """Infer partner vs direct registration from source name."""
    source = (source_name or "").lower()
    if any(keyword in source for keyword in PARTNER_KEYWORDS):
        return "Partner"
    return "Direct"


def parse_zoom_datetime(raw: str) -> Tuple[Optional[dt.datetime], Optional[str], Optional[str]]:
    """Parse Zoom timestamp into datetime, date, and time strings."""
    if not raw:
        return None, None, None
    cleaned = raw.replace("  ", " ").strip()
    for fmt in ("%b %d, %Y %I:%M %p", "%b %d, %Y %H:%M:%S", "%b %d, %Y %H:%M"):
        try:
            parsed = dt.datetime.strptime(cleaned, fmt)
            return parsed, parsed.strftime("%Y-%m-%d"), parsed.strftime("%I:%M %p")
        except ValueError:
            continue
    return None, None, raw


def build_notes(industry: str, icp_note: str, minutes: int) -> str:
    sections = []
    if industry:
        sections.append(f"Industry: {industry}")
    if icp_note:
        sections.append(f"ICP note: {icp_note}")
    sections.append(f"Minutes watched: {minutes}")
    return " | ".join(sections)


def populate_sheet(
    workbook_path: Path,
    template_name: str,
    sheet_name: str,
    product: str,
    registration_path: Path,
    attendee_path: Path,
    survey_path: Optional[Path],
    output_path: Path,
) -> None:
    metadata, registrants = load_registration_report(registration_path)
    attendance = load_attendance(attendee_path)

    event_name = metadata.get("Topic", "").strip()
    scheduled_raw = metadata.get("Scheduled Time", "")
    sched_dt, sched_date, sched_time = parse_zoom_datetime(scheduled_raw)
    display_date = sched_dt.strftime("%d/%m/%Y") if sched_dt else ""

    workbook = load_workbook(workbook_path, data_only=False)
    if sheet_name in workbook.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' already exists.")

    template_ws = workbook[template_name]
    new_ws = workbook.copy_worksheet(template_ws)
    new_ws.title = sheet_name

    # Remove placeholder rows below the template header.
    if new_ws.max_row > 20:
        new_ws.delete_rows(21, new_ws.max_row - 20)

    new_ws["B2"] = attendee_path.name
    if sched_date:
        new_ws["D9"] = sched_date
    if sched_time:
        new_ws["D10"] = sched_time

    pending_review = False

    for record in registrants:
        email = record.get("Email", "").lower()
        industry = record.get("Industry", "")

        icp_flag, icp_note = classify_icp(
            record.get("Job Title", ""), record.get("Country/Region Name", "")
        )
        if icp_flag != "ICP Confirmed":
            pending_review = True

        attendance_info = attendance.get(email, {"attended": False, "minutes": 0})
        attended_flag = "Yes" if attendance_info["attended"] else "No"
        minutes = int(attendance_info["minutes"])

        reg_time = record.get("Registration Time", "")
        try:
            reg_dt = dt.datetime.strptime(reg_time, "%b %d, %Y %H:%M:%S")
        except ValueError:
            try:
                reg_dt = dt.datetime.strptime(reg_time, "%b %d, %Y %I:%M:%S")
            except ValueError:
                reg_dt = None

        row = [
            display_date,
            product,
            event_name,
            record.get("First Name", ""),
            record.get("Last Name", ""),
            record.get("Email", ""),
            record.get("Phone", ""),
            reg_dt,
            record.get("Job Title", ""),
            "",  # Practice Name (not provided by Zoom)
            record.get("What questions do you have", ""),
            record.get("Source Name", "").strip(),
            record.get("Country/Region Name", ""),
            derive_lead_type(record.get("Source Name", "")),
            icp_flag,
            attended_flag,
            "Pending" if icp_flag != "ICP Confirmed" else "",
            build_notes(industry, icp_note, minutes),
            "",  # MSM Conversion Status
            "",  # MSM Score
            "",  # MSM Type
            "",  # MSMs Completed
            "",  # Ekwa Sales Status
            "",  # CSM Conversion Status
            "",  # CSM Type
        ]
        new_ws.append(row)

    new_ws["H1"] = "Pending Review" if pending_review else "Complete"

    # Placeholder: survey parsing can be added here when required.
    _ = survey_path  # keeps argument available for future use

    workbook.save(output_path)


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Populate Event Panel worksheet from Zoom CSV exports."
    )
    parser.add_argument("--workbook", required=True, type=Path)
    parser.add_argument("--template", default="TEMPLATE(MAKE A COPY)")
    parser.add_argument("--sheet-name", required=True)
    parser.add_argument("--product", required=True)
    parser.add_argument("--registration", required=True, type=Path)
    parser.add_argument("--attendee", required=True, type=Path)
    parser.add_argument("--survey", type=Path)
    parser.add_argument("--output", required=True, type=Path)
    args = parser.parse_args()

    populate_sheet(
        workbook_path=args.workbook,
        template_name=args.template,
        sheet_name=args.sheet_name,
        product=args.product,
        registration_path=args.registration,
        attendee_path=args.attendee,
        survey_path=args.survey,
        output_path=args.output,
    )
    print(f"Created sheet '{args.sheet_name}' in {args.output}")


if __name__ == "__main__":
    main()
