#!/usr/bin/env python3
"""Render an HTML snapshot of the Event Panel spreadsheet."""

from __future__ import annotations

import argparse
import html
from datetime import datetime
from pathlib import Path
from typing import Dict, Iterable, List, Sequence, Tuple

from openpyxl import load_workbook


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Render Event Panel data as HTML.")
    parser.add_argument("--workbook", required=True, type=Path)
    parser.add_argument("--sheet", required=True)
    parser.add_argument("--attendee-link", type=str, default="")
    parser.add_argument("--product", type=str, default="")
    parser.add_argument("--scheduled", type=str, default="")
    parser.add_argument("--output", required=True, type=Path)
    return parser.parse_args()


def iter_rows(ws, min_row: int, max_col: int) -> Iterable[Sequence]:
    for row in ws.iter_rows(min_row=min_row, min_col=1, max_col=max_col, values_only=True):
        yield row


def compute_summary(
    rows: List[Sequence],
    *,
    lead_idx: int,
    icp_idx: int,
    attendance_idx: int,
) -> Tuple[List[Tuple[str, int]], List[Tuple[str, int]]]:
    total = len(rows)

    def get_value(row: Sequence, idx: int) -> str:
        if idx < 0 or idx >= len(row):
            return ""
        value = row[idx]
        if value is None:
            return ""
        return str(value).strip()

    icp = sum(1 for r in rows if get_value(r, icp_idx) == "ICP Confirmed")
    non_icp = sum(1 for r in rows if get_value(r, icp_idx) == "Non-ICP")
    review = total - icp - non_icp

    attendees = sum(1 for r in rows if get_value(r, attendance_idx) == "Yes")
    icp_attendees = sum(
        1 for r in rows if get_value(r, icp_idx) == "ICP Confirmed" and get_value(r, attendance_idx) == "Yes"
    )
    non_icp_attendees = sum(
        1 for r in rows if get_value(r, icp_idx) == "Non-ICP" and get_value(r, attendance_idx) == "Yes"
    )

    partner_regs = sum(1 for r in rows if get_value(r, lead_idx).lower() == "partner")
    direct_regs = total - partner_regs

    summary_cards = [
        ("Total Registrations", total),
        ("ICP Registrations", icp),
        ("Non-ICP Registrations", non_icp),
        ("Needs Review", review),
        ("Total Attendees", attendees),
        ("ICP Attendees", icp_attendees),
        ("Non-ICP Attendees", non_icp_attendees),
        ("Partner Registrations", partner_regs),
        ("Direct Registrations", direct_regs),
    ]

    meta_rows = [
        ("TOTAL ICP and NON ICP REGISTRATIONS", total),
        ("TOTAL ICP REGISTRATIONS", icp),
        ("TOTAL NON ICP REGISTRATIONS", non_icp),
        ("ATTENDEES - ICP/NON ICPs", attendees),
        ("ATTENDEES - ICPs", icp_attendees),
        ("ATTENDEES - Non ICP Attendees", non_icp_attendees),
        ("TOTAL DIRECT REGISTRATIONS", direct_regs),
        ("TOTAL PARTNER REGISTRATIONS", partner_regs),
        ("Direct MSMs Booked", 0),
        ("Direct ICP MSMs Booked", 0),
        ("BDR Booked MSMs", 0),
        ("BDR ICP MSMs Booked", 0),
        ("Direct Completed MSMs", 0),
        ("BDR Completed MSMs", 0),
        ("Total ICP MSMs Booked", 0),
        ("Total ICP MSMs Completed", 0),
    ]

    return summary_cards, meta_rows


def format_datetime(value) -> str:
    if isinstance(value, datetime):
        return value.strftime("%b %d, %Y %H:%M")
    if isinstance(value, str):
        for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%b %d, %Y %H:%M:%S", "%b %d, %Y %H:%M"):
            try:
                parsed = datetime.strptime(value, fmt)
            except ValueError:
                continue
            return parsed.strftime("%b %d, %Y %H:%M")
    return str(value) if value is not None else ""


def render_html(
    *,
    headers: Sequence[str],
    rows: List[Sequence],
    summary_cards: Sequence[Tuple[str, int]],
    meta_rows: Sequence[Tuple[str, int]],
    attendee_link: str,
    scheduled_text: str,
    product: str,
    index_map: Dict[str, int],
) -> str:
    html_parts: List[str] = []
    html_parts.append(
        """<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="utf-8" />
  <title>Event Panel Snapshot</title>
  <meta name="viewport" content="width=device-width, initial-scale=1" />
  <script src="https://cdn.tailwindcss.com"></script>
  <style>
    th[data-col].active-col, td[data-col].active-col { background-color: rgba(59,130,246,0.15); }
    .copy-chip { display: inline-flex; align-items: center; gap: 0.25rem; border-radius: 999px; border: 1px solid rgba(255,255,255,0.35); background: rgba(15,23,42,0.45); color: #fff; font-size: 0.75rem; padding: 0.125rem 0.5rem; transition: background 0.15s ease; }
    .copy-chip:hover { background: rgba(15,23,42,0.65); }
  </style>
</head>
<body class="min-h-screen bg-slate-50 text-slate-900">
<div class="min-w-[1600px] px-6 py-10">
"""
    )

    html_parts.append("<header class=\"mb-8\"><h1 class=\"text-3xl font-semibold tracking-tight\">Event Panel Snapshot</h1><p class=\"mt-1 text-sm text-slate-500\">Live view generated from the panel tracking workbook.</p></header>")

    html_parts.append('<section class="grid gap-4 sm:grid-cols-2 lg:grid-cols-3 xl:grid-cols-4">')
    for title, value in summary_cards:
        html_parts.append(
            f"""
<article class="rounded-2xl bg-white p-5 shadow-sm ring-1 ring-slate-200">
  <span class="text-3xl font-semibold text-emerald-600">{value}</span>
  <p class="mt-2 text-sm font-medium text-slate-500">{html.escape(title)}</p>
</article>
"""
        )
    html_parts.append("</section>")

    html_parts.append('<section class="mt-8 grid gap-6 lg:grid-cols-[320px_minmax(0,1fr)]">')

    html_parts.append(
        '<aside class="rounded-2xl bg-white shadow-sm ring-1 ring-slate-200">'
        '<div class="border-b border-slate-200 px-5 py-4"><h2 class="text-base font-semibold">Event Health Metrics</h2></div>'
        '<dl class="divide-y divide-slate-100">'
    )
    for label, value in meta_rows:
        html_parts.append(
            f'<div class="grid grid-cols-[1fr_auto] gap-3 px-5 py-3 text-sm"><dt class="text-slate-500">{html.escape(label)}</dt><dd class="font-semibold text-slate-900">{value}</dd></div>'
        )
    html_parts.append("</dl></aside>")

    html_parts.append('<div class="rounded-2xl bg-white p-6 shadow-sm ring-1 ring-amber-200/60">')
    html_parts.append('<span class="inline-flex items-center gap-2 rounded-full bg-amber-100 px-3 py-1 text-xs font-semibold uppercase tracking-wide text-amber-700">Pending Review</span>')
    html_parts.append(
        "<p class=\"mt-4 text-sm leading-6 text-slate-600\">Data completion status is <strong>Pending</strong>. Once ICP verification is complete, update the sheet status to <em>Complete</em> before sharing with Sales.</p>"
    )
    if attendee_link:
        html_parts.append(
            f'<p class="mt-4 text-sm text-slate-600"><span class="font-semibold text-slate-800">Attendee export:</span> {html.escape(attendee_link)}</p>'
        )
    if scheduled_text:
        html_parts.append(f'<p class="text-sm text-slate-600"><span class="font-semibold text-slate-800">Scheduled:</span> {html.escape(scheduled_text)}</p>')
    if product:
        html_parts.append(f'<p class="text-sm text-slate-600"><span class="font-semibold text-slate-800">Product:</span> {html.escape(product)}</p>')
    html_parts.append("</div></section>")

    html_parts.append("<section class=\"mt-10\">")
    html_parts.append(
        '<div class="flex flex-wrap items-center gap-3 rounded-2xl bg-white px-4 py-3 shadow-sm ring-1 ring-slate-200">'
        '<label class="relative flex-1 min-w-[240px] max-w-md"><input id="searchInput" type="search" placeholder="Search table (name, email, source, notes…)" class="w-full rounded-xl border border-slate-200 bg-slate-50 px-4 py-2 text-sm text-slate-700 shadow-inner placeholder:text-slate-400 focus:border-sky-400 focus:outline-none focus:ring-2 focus:ring-sky-200" /></label>'
        '<button class="inline-flex items-center gap-2 rounded-xl bg-sky-600 px-4 py-2 text-sm font-semibold text-white shadow-sm transition hover:bg-sky-500 focus:outline-none focus:ring-2 focus:ring-sky-400 focus:ring-offset-2" id="copyTable">Copy table</button>'
        "</div>"
    )

    html_parts.append('<div class="mt-4 rounded-2xl bg-white shadow-sm ring-1 ring-slate-200"><table class="w-full min-w-[1600px] border-collapse text-sm" id="panelTable"><thead class="bg-slate-900 text-left text-xs font-semibold uppercase tracking-wide text-white"><tr>')
    for idx, heading in enumerate(headers):
        safe_heading = html.escape(heading.strip())
        html_parts.append(
            f'<th data-col="{idx}" class="sticky top-0 border-b border-slate-800 px-4 py-3 text-xs font-semibold uppercase tracking-wide"><div class="flex items-center justify-between gap-2">{safe_heading}'
            f'<button class="copy-chip col-copy" type="button" data-col="{idx}">Copy</button></div></th>'
        )
    html_parts.append("</tr></thead><tbody class=\"align-top\">")

    registration_idx = index_map.get("registration time", -1)
    lead_idx = index_map.get("lead type", -1)
    icp_idx = index_map.get("icp confirmation", -1)
    attendance_idx = index_map.get("attendance", -1)
    manager_idx = index_map.get("manager verifcation", index_map.get("manager verification", -1))

    for row in rows:
        html_parts.append("<tr>")
        for idx, value in enumerate(row):
            data_attr = f' data-col="{idx}"'
            if idx == registration_idx:
                cell_value = format_datetime(value)
            elif idx == lead_idx:
                text = (value or "").strip()
                if text.lower() == "partner":
                    cell_value = '<span class="inline-flex items-center rounded-full bg-sky-100 px-2.5 py-1 text-xs font-semibold text-sky-700">Partner</span>'
                elif text.lower() == "direct":
                    cell_value = '<span class="inline-flex items-center rounded-full bg-amber-100 px-2.5 py-1 text-xs font-semibold text-amber-700">Direct</span>'
                else:
                    cell_value = html.escape(text)
            elif idx == icp_idx:
                status = (value or "").strip()
                if status == "ICP Confirmed":
                    cell_value = '<span class="inline-flex items-center rounded-full bg-emerald-100 px-2.5 py-1 text-xs font-semibold text-emerald-700">ICP Confirmed</span>'
                elif status == "Non-ICP":
                    cell_value = '<span class="inline-flex items-center rounded-full bg-rose-100 px-2.5 py-1 text-xs font-semibold text-rose-700">Non-ICP</span>'
                elif status:
                    cell_value = '<span class="inline-flex items-center rounded-full bg-amber-100 px-2.5 py-1 text-xs font-semibold text-amber-700">Needs Review</span>'
                else:
                    cell_value = ""
            elif idx == attendance_idx:
                attendance = (value or "").strip()
                if attendance == "Yes":
                    cell_value = '<span class="inline-flex items-center rounded-full bg-emerald-100 px-2.5 py-1 text-xs font-semibold text-emerald-700">Yes</span>'
                elif attendance == "No":
                    cell_value = '<span class="inline-flex items-center rounded-full bg-rose-100 px-2.5 py-1 text-xs font-semibold text-rose-700">No</span>'
                else:
                    cell_value = ""
            elif idx == manager_idx:
                verification = (value or "").strip()
                if verification == "Pending":
                    cell_value = '<span class="text-lg text-slate-400" title="Pending verification">&#x2610;</span>'
                elif verification:
                    cell_value = '<span class="text-lg text-emerald-600" title="Verified">&#x2611;</span>'
                else:
                    cell_value = ""
            else:
                if value is None:
                    cell_value = ""
                else:
                    if isinstance(value, datetime):
                        cell_value = html.escape(format_datetime(value))
                    else:
                        text_value = str(value)
                        cell_value = html.escape(text_value)
            cell_classes = "border-b border-slate-100 px-4 py-3 text-sm text-slate-700" + (" whitespace-pre-line" if idx == 17 else " whitespace-nowrap")
            html_parts.append(f"<td{data_attr} class=\"{cell_classes}\">{cell_value}</td>")
        html_parts.append("</tr>")
    html_parts.append("</tbody></table></div>")
    html_parts.append("</section>")

    html_parts.append(
        """<footer class="mt-10 text-xs text-slate-400">Generated from Zoom exports and ICP guidelines. Provide new CSVs to refresh this dashboard.</footer>
</div>
<script>
  const searchInput = document.getElementById("searchInput");
  const table = document.getElementById("panelTable");

  const highlightColumn = (col) => {
    for (const cell of table.querySelectorAll("[data-col]")) {
      cell.classList.toggle("active-col", Number(cell.dataset.col) === col);
    }
    for (const header of table.tHead.rows[0].cells) {
      header.classList.toggle("active-col", Number(header.dataset.col) === col);
    }
  };

  const copyColumn = (col) => {
    const visibleRows = Array.from(table.tBodies[0].rows).filter(
      (row) => row.style.display !== "none"
    );
    const values = visibleRows.map((row) => row.cells[col].innerText.trim());
    const payload = values.join("\\n");

    if (navigator.clipboard && navigator.clipboard.writeText) {
      navigator.clipboard.writeText(payload).catch(() => {
        const textarea = document.createElement("textarea");
        textarea.value = payload;
        document.body.appendChild(textarea);
        textarea.select();
        document.execCommand("copy");
        document.body.removeChild(textarea);
      });
    } else {
      const textarea = document.createElement("textarea");
      textarea.value = payload;
      document.body.appendChild(textarea);
      textarea.select();
      document.execCommand("copy");
      document.body.removeChild(textarea);
    }
  };

  searchInput.addEventListener("input", () => {
    const query = searchInput.value.toLowerCase();
    for (const row of table.tBodies[0].rows) {
      const text = row.innerText.toLowerCase();
      row.style.display = text.includes(query) ? "" : "none";
    }
  });

  for (const header of table.tHead.rows[0].cells) {
    header.addEventListener("mouseenter", () => highlightColumn(Number(header.dataset.col)));
    header.addEventListener("mouseleave", () => highlightColumn(-1));
  }

  table.addEventListener("mouseleave", () => highlightColumn(-1));

  for (const button of document.querySelectorAll(".col-copy")) {
    button.addEventListener("click", (event) => {
      event.stopPropagation();
      const col = Number(button.dataset.col);
      highlightColumn(col);
      copyColumn(col);
    });
  }

  document.getElementById("copyTable").addEventListener("click", () => {
    const selection = window.getSelection();
    const range = document.createRange();
    range.selectNode(table);
    selection.removeAllRanges();
    selection.addRange(range);
    try {
      document.execCommand("copy");
    } catch (err) {
      console.warn("Copy failed", err);
    }
    selection.removeAllRanges();
  });
</script>
</body>
</html>
"""
    )

    return "".join(html_parts)


def main() -> None:
    args = parse_args()

    wb = load_workbook(args.workbook, data_only=True)
    ws = wb[args.sheet]

    max_col = ws.max_column
    header_row = list(iter_rows(ws, min_row=20, max_col=max_col))[0]
    headers_raw = list(header_row)
    while headers_raw and (headers_raw[-1] is None or str(headers_raw[-1]).strip() == ""):
        headers_raw.pop()
    headers = [str(h).strip() if h else "" for h in headers_raw]
    effective_cols = len(headers)

    data_rows: List[Sequence] = []
    for row in iter_rows(ws, min_row=21, max_col=max_col):
        sliced = list(row[:effective_cols])
        if not any(sliced):
            continue
        data_rows.append(sliced)

    index_map: Dict[str, int] = {}
    for idx, header in enumerate(headers):
        key = header.strip().lower()
        if key:
            index_map[key] = idx

    lead_idx = index_map.get("lead type", -1)
    icp_idx = index_map.get("icp confirmation", -1)
    attendance_idx = index_map.get("attendance", -1)

    summary_cards, meta_rows = compute_summary(
        data_rows,
        lead_idx=lead_idx,
        icp_idx=icp_idx,
        attendance_idx=attendance_idx,
    )

    html_output = render_html(
        headers=headers,
        rows=data_rows,
        summary_cards=summary_cards,
        meta_rows=meta_rows,
        attendee_link=args.attendee_link,
        scheduled_text=args.scheduled,
        product=args.product,
        index_map=index_map,
    )

    args.output.write_text(html_output, encoding="utf-8")
    print(f"Created {args.output}")


if __name__ == "__main__":
    main()
